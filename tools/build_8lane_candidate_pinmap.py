from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
ACTIVE_XDC = ROOT / "TFDU_VFIR_Client_Array" / "TFDU_VFIR_Client.srcs" / "constrs_1" / "new" / "PORT1.xdc"
CANDIDATE_XDC = ROOT / "TFDU_VFIR_Client_Array" / "TFDU_VFIR_Client.srcs" / "constrs_1" / "new" / "target_ir_array_8lane_candidate.xdc"
EXPECTED_CONSTRAINT_SHA256 = "CFF1A17EE77BBAF90080CF4F97E5920E961AEFAE3B6752F080E35FCF4D4B1F11"
KNOWN_EXCLUDED_PINS = {
    "D19": "previous B_RX1 location was bypassed during hardware debug; keep out of automatic 8-lane candidate until manually cleared",
}
SIGNALS = ["MODE", "RX", "SD", "TX"]
PORT_BASE = {
    ("A", "MODE"): "ir_mode_out_0",
    ("A", "RX"): "ir_rx_in_0",
    ("A", "SD"): "ir_sd_0",
    ("A", "TX"): "ir_tx_out_0",
    ("B", "MODE"): "loop_mode_b0",
    ("B", "RX"): "loop_rx_b0",
    ("B", "SD"): "loop_sd_b0",
    ("B", "TX"): "loop_tx_b0",
}


@dataclass
class AvailablePin:
    package_pin: str
    board_signal: str
    connector_group: str
    index: int
    polarity: str


@dataclass
class CandidateAssignment:
    endpoint: str
    lane: int
    signal: str
    port: str
    package_pin: str
    board_signal: str
    connector_group: str
    origin: str
    note: str


def sha256(path: Path | None) -> str:
    if path is None or not path.exists():
        return "MISSING"
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def read_text(path: Path | None) -> str:
    if path is None or not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def rel(path: Path | None) -> str:
    if path is None:
        return ""
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path)


def find_hard_constraint() -> Path | None:
    for path in ROOT.glob("*.txt"):
        if sha256(path) == EXPECTED_CONSTRAINT_SHA256:
            return path
    return None


def find_pin_workbook() -> Path:
    candidates = sorted((ROOT / "hardware" / "01_SCH").glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No AX7010 pin workbook found under hardware/01_SCH")
    return candidates[0]


def parse_tfdu_port(port: str) -> tuple[str, int, str] | None:
    reverse = {base: (endpoint, signal) for (endpoint, signal), base in PORT_BASE.items()}
    match = re.fullmatch(r"([A-Za-z0-9_]+)\[(\d+)\]", port)
    if match is None:
        return None
    base = match.group(1)
    if base not in reverse:
        return None
    endpoint, signal = reverse[base]
    return endpoint, int(match.group(2)), signal


def parse_active_assignments(path: Path) -> dict[tuple[str, int, str], CandidateAssignment]:
    text = read_text(path)
    pin_by_port: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        match = re.search(r"set_property\s+PACKAGE_PIN\s+(\S+)\s+\[get_ports\s+\{?([^}\]]+\][^}]*)\}?\]", line)
        if match:
            pin_by_port[match.group(2).strip()] = match.group(1).strip().strip('"')

    active: dict[tuple[str, int, str], CandidateAssignment] = {}
    for port, pin in pin_by_port.items():
        parsed = parse_tfdu_port(port)
        if parsed is None:
            continue
        endpoint, lane, signal = parsed
        active[(endpoint, lane, signal)] = CandidateAssignment(
            endpoint=endpoint,
            lane=lane,
            signal=signal,
            port=port,
            package_pin=pin,
            board_signal="",
            connector_group="",
            origin="active_PORT1",
            note="Preserved from current 2-lane hardware XDC.",
        )
    return active


def load_available_pins(workbook: Path) -> list[AvailablePin]:
    wb = load_workbook(workbook, data_only=True)
    ws = wb.worksheets[0]
    pins: list[AvailablePin] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        package_pin = str(row[0] or "").strip().upper()
        board_signal = str(row[1] or "").strip()
        match = re.fullmatch(r"IO([12])_(\d+)([NP])", board_signal)
        if not match:
            continue
        pins.append(
            AvailablePin(
                package_pin=package_pin,
                board_signal=board_signal,
                connector_group=f"IO{match.group(1)}",
                index=int(match.group(2)),
                polarity=match.group(3),
            )
        )
    return sorted(pins, key=lambda p: (p.connector_group, p.index, 0 if p.polarity == "P" else 1, p.package_pin))


def enrich_active(active: dict[tuple[str, int, str], CandidateAssignment], available: list[AvailablePin]) -> None:
    by_pin = {pin.package_pin: pin for pin in available}
    for assignment in active.values():
        pin = by_pin.get(assignment.package_pin)
        if pin is not None:
            assignment.board_signal = pin.board_signal
            assignment.connector_group = pin.connector_group


def desired_keys() -> list[tuple[str, int, str]]:
    return [(endpoint, lane, signal) for endpoint in ["A", "B"] for lane in range(8) for signal in SIGNALS]


def make_port(endpoint: str, lane: int, signal: str) -> str:
    return f"{PORT_BASE[(endpoint, signal)]}[{lane}]"


def build_candidate(active: dict[tuple[str, int, str], CandidateAssignment], available: list[AvailablePin]) -> list[CandidateAssignment]:
    enrich_active(active, available)
    used_pins = {assignment.package_pin for assignment in active.values()}
    free = [
        pin
        for pin in available
        if pin.package_pin not in used_pins and pin.package_pin not in KNOWN_EXCLUDED_PINS
    ]
    free_iter = iter(free)

    candidate: list[CandidateAssignment] = []
    for key in desired_keys():
        endpoint, lane, signal = key
        if key in active:
            candidate.append(active[key])
            continue
        try:
            pin = next(free_iter)
        except StopIteration as exc:
            raise RuntimeError("Not enough IO1/IO2 pins remain to build an 8-lane A/B candidate pinmap") from exc
        candidate.append(
            CandidateAssignment(
                endpoint=endpoint,
                lane=lane,
                signal=signal,
                port=make_port(endpoint, lane, signal),
                package_pin=pin.package_pin,
                board_signal=pin.board_signal,
                connector_group=pin.connector_group,
                origin="candidate_auto",
                note="Auto-assigned from currently unused IO1/IO2 expansion pins; review against physical connectors before hardware use.",
            )
        )
    return candidate


def validate_candidate(candidate: list[CandidateAssignment]) -> None:
    if len(candidate) != 64:
        raise RuntimeError(f"Expected 64 TFDU signal assignments for 8-lane A/B candidate, got {len(candidate)}")
    keys = {(item.endpoint, item.lane, item.signal) for item in candidate}
    expected = set(desired_keys())
    if keys != expected:
        missing = sorted(expected - keys)
        extra = sorted(keys - expected)
        raise RuntimeError(f"Candidate key mismatch missing={missing} extra={extra}")
    pins: dict[str, list[str]] = {}
    for item in candidate:
        pins.setdefault(item.package_pin, []).append(item.port)
    duplicates = {pin: ports for pin, ports in pins.items() if len(ports) > 1}
    if duplicates:
        raise RuntimeError(f"Duplicate package pins in candidate: {duplicates}")
    excluded_used = sorted(pin for pin in KNOWN_EXCLUDED_PINS if pin in pins)
    if excluded_used:
        raise RuntimeError(f"Known excluded pins used unexpectedly: {excluded_used}")


def write_candidate_xdc(path: Path, candidate: list[CandidateAssignment], generated: str) -> None:
    lines = [
        "###############################################################################",
        "# 8-lane TFDU candidate pin map",
        "#",
        f"# Generated: {generated}",
        "# Source: hardware/01_SCH AX7010/AX7020 pin workbook + current PORT1.xdc",
        "#",
        "# This is a candidate for review, not proven hardware acceptance.",
        "# Do not promote it to the active constraints until the physical connector",
        "# mapping, IO bank voltage, shutdown coverage, and TFDU wiring are reviewed.",
        "###############################################################################",
        "",
    ]
    for endpoint in ["A", "B"]:
        lines.extend([f"# Endpoint {endpoint}", ""])
        for lane in range(8):
            lines.append(f"# Endpoint {endpoint}, lane {lane}")
            for signal in SIGNALS:
                item = next(a for a in candidate if a.endpoint == endpoint and a.lane == lane and a.signal == signal)
                lines.append(f"# {item.port} -> {item.board_signal} / {item.package_pin} / {item.origin}")
                lines.append(f"set_property PACKAGE_PIN {item.package_pin} [get_ports {{{item.port}}}]")
                lines.append(f"set_property IOSTANDARD LVCMOS33 [get_ports {{{item.port}}}]")
            lines.append("")
    lines.extend(
        [
            "###############################################################################",
            "# Known excluded automatic pin choices",
            "###############################################################################",
        ]
    )
    for pin, reason in sorted(KNOWN_EXCLUDED_PINS.items()):
        lines.append(f"# {pin}: {reason}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_reports(candidate: list[CandidateAssignment], workbook: Path, generated: str) -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    md_path = REPORTS / "8lane_candidate_pinmap_current.md"
    json_path = REPORTS / "8lane_candidate_pinmap_current.json"
    csv_path = REPORTS / "8lane_candidate_pinmap_current.csv"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(candidate[0]).keys()))
        writer.writeheader()
        for item in candidate:
            writer.writerow(asdict(item))

    rows = [
        [
            item.endpoint,
            str(item.lane),
            item.signal,
            item.port,
            item.package_pin,
            item.board_signal,
            item.connector_group,
            item.origin,
            item.note,
        ]
        for item in candidate
    ]
    md = [
        "# 8-Lane Candidate Pin Map",
        "",
        f"Generated: {generated}",
        "",
        "## Verdict",
        "",
        "- Overall: `CANDIDATE_PINMAP_GENERATED_REVIEW_REQUIRED`",
        "- Candidate TFDU assignments: `64/64`",
        "- Candidate lanes: `8 A-endpoint + 8 board-internal B-endpoint`",
        "- This is not real hardware acceptance and is not automatically promoted to the active XDC.",
        "",
        "## Sources",
        "",
        f"- Pin workbook: `{rel(workbook)}`",
        f"- Current active XDC preserved for existing lanes: `{rel(ACTIVE_XDC)}`",
        f"- Candidate XDC written to: `{rel(CANDIDATE_XDC)}`",
        "",
        "## Safety",
        "",
        "- No FPGA was programmed.",
        "- No UART was written.",
        "- No TFDU was driven.",
        "- Excluded automatic pin choices are kept out of the candidate.",
        "",
        "## Candidate Table",
        "",
        "| endpoint | lane | signal | port | package_pin | board_signal | group | origin | note |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in rows:
        md.append("| " + " | ".join(cell.replace("|", "/") for cell in row) + " |")
    md.extend(
        [
            "",
            "## Excluded Pins",
            "",
        ]
    )
    for pin, reason in sorted(KNOWN_EXCLUDED_PINS.items()):
        md.append(f"- `{pin}`: {reason}")
    md.extend(
        [
            "",
            "RF_COMM_8LANE_CANDIDATE_PINMAP overall=CANDIDATE_PINMAP_GENERATED_REVIEW_REQUIRED assignments=64",
        ]
    )
    md_path.write_text("\n".join(md) + "\n", encoding="utf-8")

    payload = {
        "generated": generated,
        "overall": "CANDIDATE_PINMAP_GENERATED_REVIEW_REQUIRED",
        "hard_constraint_sha256": sha256(find_hard_constraint()),
        "pin_workbook": rel(workbook),
        "active_xdc": rel(ACTIVE_XDC),
        "candidate_xdc": rel(CANDIDATE_XDC),
        "known_excluded_pins": KNOWN_EXCLUDED_PINS,
        "no_hardware_programming": True,
        "no_uart_write": True,
        "no_tfdu_drive": True,
        "assignments": [asdict(item) for item in candidate],
    }
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> int:
    constraint = find_hard_constraint()
    if constraint is None:
        raise RuntimeError("Hard constraint file was not found by expected hash")
    workbook = find_pin_workbook()
    available = load_available_pins(workbook)
    active = parse_active_assignments(ACTIVE_XDC)
    candidate = build_candidate(active, available)
    validate_candidate(candidate)
    generated = datetime.now().isoformat(timespec="seconds")
    write_candidate_xdc(CANDIDATE_XDC, candidate, generated)
    write_reports(candidate, workbook, generated)

    print(f"WROTE_XDC={CANDIDATE_XDC}")
    print(f"WROTE_MARKDOWN={REPORTS / '8lane_candidate_pinmap_current.md'}")
    print(f"WROTE_JSON={REPORTS / '8lane_candidate_pinmap_current.json'}")
    print(f"WROTE_CSV={REPORTS / '8lane_candidate_pinmap_current.csv'}")
    print("RF_COMM_8LANE_CANDIDATE_PINMAP overall=CANDIDATE_PINMAP_GENERATED_REVIEW_REQUIRED assignments=64")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
